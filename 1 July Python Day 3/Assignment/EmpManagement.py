import csv
import datetime

# import openpyxl for handling Excel files
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# import reportlab for PDF generation
from reportlab.pdfgen import canvas
from Employee import Employee

# EmployeeManagementSystem class
class EmployeeManagementSystem:

    # Constructor to initialize EmployeeManagementSystem object
    def __init__(self):
        self.employees = []

    # Method to add employee to the system
    def add_employee(self, employee):
        self.employees.append(employee)

    # Method to save employee data to text file
    def save_to_text(self, file_name):
        with open(file_name, 'w') as file:
            for emp in self.employees:
                file.write(f"Name: {emp.name}\nDOB: {emp.dob}\nPhone: {emp.phone}\nEmail: {emp.email}\nAge: {emp.calculate_age()}\n\n")
        print(f"Data saved to {file_name}")

    # Method to save employee data to Excel file
    def save_to_excel(self, file_name):

        # Create a new Excel workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Set the title of the worksheet
        ws.title = "Employees"
        # Add headers to the worksheet
        headers = ['Name', 'DOB', 'Phone', 'Email', 'Age']

        ws.append(headers)

        # Add employee data to the worksheet
        for emp in self.employees:
            ws.append([emp.name, emp.dob, emp.phone, emp.email, emp.calculate_age()])
        
        # Auto-size columns based on content
        for col in range(1, 6):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].auto_size = True
        
        # Save the workbook to the specified file
        wb.save(file_name)
        print(f"Data saved to {file_name}")

    # Method to save employee data to PDF file
    def save_to_pdf(self, file_name):

        # Create a new PDF document
        c = canvas.Canvas(file_name)
        y = 750

        # Add employee data to the PDF document
        for emp in self.employees:
            c.drawString(100, y, f"Name: {emp.name}")
            c.drawString(100, y - 20, f"DOB: {emp.dob}")
            c.drawString(100, y - 40, f"Phone: {emp.phone}")
            c.drawString(100, y - 60, f"Email: {emp.email}")
            c.drawString(100, y - 80, f"Age: {emp.calculate_age()}")
            y -= 100
        c.save()
        print(f"Data saved to {file_name}")

    def bulk_read_from_excel(self, file_name):
        # Implementation for bulk reading from Excel can be added here
        pass

# Main program execution
def main():
    emp_management = EmployeeManagementSystem()

    while True:
        print("\nEmployee Management System")
        print("1. Add Employee")
        print("\nWant to Save the data to file?")
        print("2. Save to Text")
        print("3. Save to Excel")
        print("4. Save to PDF")
        print("5. Exit")

        choice = input("Enter your choice: ")

        if choice == '1':
            name = input("Enter Name: ")
            dob = input("Enter DOB (YYYY-MM-DD): ")
            phone = input("Enter Phone: ")
            email = input("Enter Email: ")

            # Validate input and create employee object
            # Example validation: check if DOB is in correct format
            try:
                datetime.datetime.strptime(dob, '%Y-%m-%d')
            except ValueError:
                print("Incorrect date format, should be YYYY-MM-DD")
                continue
            
            emp = Employee(name, dob, phone, email)
            emp_management.add_employee(emp)
            print("Employee added successfully!")

        elif choice == '2':
            file_name = input("Enter filename to save (e.g., employees.txt): ")
            emp_management.save_to_text(file_name)

        elif choice == '3':
            file_name = input("Enter filename to save (e.g., employees.xlsx): ")
            emp_management.save_to_excel(file_name)

        elif choice == '4':
            file_name = input("Enter filename to save (e.g., employees.pdf): ")
            emp_management.save_to_pdf(file_name)

        elif choice == '5':
            print("Exiting program...")
            break

        else:
            print("Invalid choice. Please choose again.")

if __name__ == "__main__":
    main()
